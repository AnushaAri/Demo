"""Program to compare two text files and write score in excel file"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

DEFAULT_ANS_FILE = "ans.txt"
DEFAULT_OUT_FILENAME = "Marksheet.xlsx"
DEFAULT_SHEET_NAME = "Subject"
DEFAULT_NO_HEAD_PARA = 4
DEFAULT_HEAD_PARA = ["S.No", "Roll No", "Score", "Status"]
DEFAULT_PASS_MARK = 15

def input_files():
    """Taking input file names"""
    ans_file = input("Enter answer sheet name (default:ans.txt) ")

    if ans_file == '':
        ans_file = DEFAULT_ANS_FILE

    number = input("Enter no of files: ")
    sequence = input("Do you have sequential numbered files? (Y/N): ")
    file_list = []

    while True:
        if sequence == ("Y" and "y"):
            start = int(input("Enter starting number "))
            last = start + (int(number))
            for step1 in range(start, last):
                ext = (str(step1)) + ".txt"
                file_list.append(ext)
            break
        elif sequence == ("N" and "n"):
            print("Enter the file names: ")
            file_list = [input() for step in range(int(number))]
            break
        else:
            sequence = input("(Y?N) : ")

    return number, ans_file, file_list

def sheet_parameters():
    """Taking input for sheet parameters from user"""
    output_filename = input("Enter excel filename (Default: Marksheet.xlsx) ")

    if os.path.exists(output_filename):
        prom = input("File already exists, do you want to overwrite it? (Y/N): ")
        while True:
            if prom == ("Y" and "y"):
                os.remove(output_filename)
                break
            elif prom == ("N" and "n"):
                output_filename = input("Enter new excel filename: ")
                break
            else:
                prom = input("(Y/N) : ")

    if output_filename == '':
        output_filename = DEFAULT_OUT_FILENAME
        if os.path.exists(output_filename):
            os.remove(output_filename)

    sheet_name = input("Enter sheet name (Default: Subject)  ")

    if sheet_name == '':
        sheet_name = DEFAULT_SHEET_NAME

    print("Default no of headers - 3 & names - S.No, Roll No, Score, Status")
    no_header_para = input("Enter no of header parameters ")

    if no_header_para == '':
        no_header_para = DEFAULT_NO_HEAD_PARA
        header_parameter = list(DEFAULT_HEAD_PARA)
    else:
        print("Enter names of header parameter: ")
        header_parameter = [input() for i in range(int(no_header_para))]

    cut_off = input("Enter the cut-off marks (Default: 15)")

    if cut_off == '':
        cut_off = DEFAULT_PASS_MARK

    return output_filename, sheet_name, header_parameter, cut_off, no_header_para

def create_workbook(number, output_filename, sheet_name, header_parameter):
    """Creating excel workbook using openpyxl"""
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = sheet_name
    rows = 1
    cols = 1

    for num in header_parameter:
        cell = work_sheet.cell(row=rows, column=cols)
        cell.value = num
        cell.font = cell.font.copy(bold=True)
        cols += 1

    cols = 1
    rows += 1
    end = int(number) + 1

    for num1 in range(1, end):
        work_sheet.cell(row=rows, column=cols).value = num1
        rows += 1

    work_book.save(output_filename)

def compare_files(files, ans_file):
    """comparing two text files line by line"""
    count = 0
    fileptr1 = open(files, 'r')
    fileptr2 = open(ans_file, 'r')

    for line1 in fileptr1:
        for line2 in fileptr2:
            if line1 == line2:
                count += 1
            break

    fileptr2.close()
    fileptr2.close()
    return count

def multi_files(file_list, ans_file, cut_off):
    """Looping comparison for multiple input files"""
    score = []
    status = []

    for step1 in file_list:
        files = step1
        score1 = compare_files(files, ans_file)
        score.append(score1)

    for com in score:
        if com >= (int(cut_off)):
            status.append("PASS")
        else:
            status.append("FAIL")

    return score, status

def write_file(score, file_list, output_filename, status):
    """Writing score in the excel sheet"""
    work_book = load_workbook(output_filename)
    work_sheet = work_book.active
    rows = 2
    cols = 2
    pattern = [search.split(".")[0] for search in file_list]
    for i in pattern:
        work_sheet.cell(row=rows, column=cols).value = i
        rows += 1

    rows = 2
    cols += 1
    for j in score:
        work_sheet.cell(row=rows, column=cols).value = j
        rows += 1

    rows = 2
    cols += 1
    for k in status:
        work_sheet.cell(row=rows, column=cols).value = k
        rows += 1

    work_book.save(output_filename)

def text_alignment(output_filename, no_header_para, number):
    """Aligning cells at the centre"""
    work_book = load_workbook(output_filename)
    work_sheet = work_book.active
    rows = 1
    cols = 1
    row_range = (int(number)) + 2

    for row_align in range(row_range):
        for col_align in range(int(no_header_para)):
            work_sheet.cell(row=rows, column=cols).alignment = Alignment(horizontal="center")
            cols += 1
        cols = 1
        rows += 1

    work_book.save(output_filename)

NUM, ANS_SHEET, FILE_LIST = input_files()
OUT_FILENAME, SHEETNAME, HEAD_PARA, PASS_MARK, NO_HEAD_PARA = sheet_parameters()
create_workbook(NUM, OUT_FILENAME, SHEETNAME, HEAD_PARA)
SCORES, STATUS = multi_files(FILE_LIST, ANS_SHEET, PASS_MARK)
write_file(SCORES, FILE_LIST, OUT_FILENAME, STATUS)
text_alignment(OUT_FILENAME, NO_HEAD_PARA, NUM)


